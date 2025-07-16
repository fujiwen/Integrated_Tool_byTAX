import pandas as pd
import warnings
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.worksheet.page import PageMargins
from openpyxl.worksheet.properties import WorksheetProperties, PageSetupProperties
from datetime import datetime
import os
from tkinter import *
from tkinter import ttk, filedialog, messagebox
import threading
import shutil  # 新增导入
import sys
import subprocess

class BldBuyApp:
    def __init__(self, root):
        self.root = root
        self.root.title("供应商供货明细表工具byTAX")
        
        # 设置窗口大小并居中
        self.set_window_geometry(580, 650)
        
        # 使窗口前台显示
        self.bring_to_front()
        
        # 检查时间验证
        if not self.check_expiration():
            messagebox.showerror("错误", "Dll注册失败，请联系开发者Cayman 13111986898")
            self.root.destroy()
            return
            
        # 定义期望的表头字段
        self.expected_headers = [
            "收货日期", "订单号", "商品名称", "实收数量", "基本单位",
            "单价(结算)", "小计金额(结算)", "税额(结算)", "小计价税(结算)", "部门",
            "税率", "供应商/备用金报销账户","商品分类"
        ]
        
        # 创建主框架
        self.main_frame = ttk.Frame(root, padding="10")
        self.main_frame.pack(fill=BOTH, expand=True)
        
        # 创建控制面板
        self.create_control_panel()
        
        # 创建日志显示区域
        self.create_log_area()
        
        # 初始化状态
        self.processing = False
        
        # 创建开发者信息标签
        self.create_developer_label()
        
    def set_window_geometry(self, width, height):
        """设置窗口大小并居中"""
        screen_width = self.root.winfo_screenwidth()
        screen_height = self.root.winfo_screenheight()
        
        x = (screen_width - width) // 2
        y = (screen_height - height) // 2
        
        self.root.geometry(f"{width}x{height}+{x}+{y}")
        
    def check_expiration(self):
        """检查时间是否到期"""
        current_date = datetime.now()
        expiration_date = datetime(2099, 12, 31)  # 年底到期
        
        return current_date <= expiration_date
        
    def create_control_panel(self):
        control_frame = ttk.LabelFrame(self.main_frame, text="请选择[收货单商品明细]报表", padding="10")
        control_frame.pack(fill=X, pady=5)
        
        # 修改为选择文件按钮
        self.file_frame = ttk.Frame(control_frame)
        self.file_frame.pack(fill=X, pady=5)
        
        ttk.Label(self.file_frame, text="选择文件:").pack(side=LEFT)
        self.input_file_var = StringVar()
        ttk.Entry(self.file_frame, textvariable=self.input_file_var, width=40).pack(side=LEFT, padx=5)
        ttk.Button(self.file_frame, text="浏览...", command=self.select_input_file).pack(side=LEFT)
        
        # 处理按钮
        self.process_btn = ttk.Button(control_frame, text="开始处理", command=self.start_processing)
        self.process_btn.pack(pady=10)
        
        # 进度条
        self.progress = ttk.Progressbar(control_frame, orient=HORIZONTAL, mode='determinate')
        self.progress.pack(fill=X, pady=5)
        
    def create_log_area(self):
        log_frame = ttk.LabelFrame(self.main_frame, text="处理日志", padding="10")
        log_frame.pack(fill=BOTH, expand=True)
        
        self.log_text = Text(log_frame, wrap=WORD, state=DISABLED)
        scrollbar = ttk.Scrollbar(log_frame, command=self.log_text.yview)
        self.log_text.configure(yscrollcommand=scrollbar.set)
        
        scrollbar.pack(side=RIGHT, fill=Y)
        self.log_text.pack(fill=BOTH, expand=True)
        
    def select_input_file(self):
        filetypes = [("Excel files", "*.xlsx *.xls")]
        file_paths = filedialog.askopenfilenames(filetypes=filetypes)
        if file_paths:
            self.input_file_var.set("\n".join(file_paths))  # 用换行符分隔多个文件路径
            
    def log_message(self, message):
        """修改后的日志记录方法"""
        # 将消息添加到日志列表
        self.log_messages.append(message)
        
        # 实时显示日志
        self.log_text.config(state=NORMAL)
        if message.startswith("警告："):
            self.log_text.tag_config("warning", foreground="red")
            self.log_text.insert(END, message + "\n", "warning")
        else:
            self.log_text.insert(END, message + "\n")
        self.log_text.see(END)
        self.log_text.config(state=DISABLED)
        
    def start_processing(self):
        if self.processing:
            return
            
        self.processing = True
        self.process_btn.config(state=DISABLED)
        self.log_text.delete(1.0, END)
        self.progress['value'] = 0
        
        # 使用线程处理，避免界面卡顿
        threading.Thread(target=self.process_files, daemon=True).start()
        
    def preprocess_excel(self, file_path):
        """预处理Excel文件，自动搜索表头位置"""
        # 自动搜索表头位置
        header_row = self.find_header_row(file_path)
        
        # 使用找到的表头行读取数据
        df = pd.read_excel(file_path, skiprows=header_row)
        
        # 添加需要保留的退货相关列，排除N-R列数据
        required_columns = self.expected_headers + ['退货', '合计退货数量', '退货合计金额(结算)', '退货合计税额(结算)', '退货合计价税(结算)']
        
        # 过滤并重新排列列，排除N-R列的数据
        exclude_columns = df.iloc[:, 13:17].columns.tolist()  # N-R列的索引是13-17
        df = df.drop(columns=exclude_columns, errors='ignore')
        
        # 过滤并保留所需列
        df_filtered = df.reindex(columns=[col if col != '单位' else '基本单位' for col in required_columns if col in df.columns or col == '基本单位'])
        
        # 处理收货日期，去掉时间部分
        if '收货日期' in df_filtered.columns:
            df_filtered['收货日期'] = pd.to_datetime(df_filtered['收货日期'], errors='coerce').dt.strftime('%Y-%m-%d')
        
        return df_filtered.dropna(how='all')
        
    def find_header_row(self, file_path):
        """自动搜索Excel文件中的表头行"""
        # 最大搜索行数
        max_rows = 50
        
        # 读取前max_rows行来查找表头
        sample_df = pd.read_excel(file_path, nrows=max_rows, header=None)
        
        # 定义匹配度阈值（至少需要匹配的表头数量）
        min_match_threshold = 3
        
        # 遍历每一行，检查是否包含足够多的预期表头
        for i in range(max_rows):
            row = sample_df.iloc[i].astype(str)
            # 计算当前行与预期表头的匹配数量
            matches = sum(1 for header in self.expected_headers if any(header in str(cell) for cell in row))
            
            # 如果匹配数量超过阈值，认为找到了表头行
            if matches >= min_match_threshold:
                self.log_message(f"找到表头行: 第{i+1}行，匹配度: {matches}/{len(self.expected_headers)}")
                return i
        
        # 如果没有找到，使用默认值
        self.log_message("未找到表头行，使用默认值(35)")
        return 35
        
    def process_files(self):
        try:
            # 初始化日志列表
            self.log_messages = []
            
            input_files = self.input_file_var.get().split("\n")
            if not input_files or not input_files[0]:
                self.log_message("请先选择要处理的Excel文件")
                return
                
            output_folder = "export"
            archive_folder = "archive"
            
            # 确保文件夹存在
            for folder in [output_folder, archive_folder]:
                if not os.path.exists(folder):
                    os.makedirs(folder)
                    self.log_message(f"创建文件夹: {folder}")
            
            # 计算总文件数
            total_files = len([f for f in input_files if f])
            processed_files = 0
            
            # 处理每个文件
            for input_file in input_files:
                if not input_file:  # 跳过空路径
                    continue
                    
                self.log_message(f"\n正在处理文件: {os.path.basename(input_file)}")
                
                try:
                    # 读取config.txt获取标题信息
                    # 获取程序运行路径
                    if getattr(sys, 'frozen', False):
                        # 如果是打包后的exe运行
                        application_path = os.path.dirname(sys.executable)
                    else:
                        # 如果是python脚本运行
                        application_path = os.path.dirname(os.path.abspath(__file__))
                    config_file = os.path.join(application_path, 'config.txt')
                    if os.path.exists(config_file):
                        with open(config_file, 'r', encoding='utf-8') as f:
                            config_lines = f.readlines()
                        
                        # 解析config.txt中的信息
                        hotelname = ''
                        sheet_title = ''
                        for line in config_lines:
                            if line.startswith('hotelname:'):
                                hotelname = line.split(':', 1)[1].strip()
                            elif line.startswith('Sheet_tittle:'):
                                sheet_title = line.split(':', 1)[1].strip()
                        
                        # 创建标题行
                        header_rows = [
                            [''] * 13,
                            [''] * 5 + [hotelname] + [''] * 7,
                            [''] * 5 + [sheet_title] + [''] * 7,
                            [''] * 13,
                            [''] * 13
                        ]
                    else:
                        header_rows = []
                        self.log_message("警告：未找到config.txt文件,将会导致对帐单标题错误")
                    
                    df_filtered = self.preprocess_excel(input_file)
                    
                    # 检查表头
                    missing_columns = set(self.expected_headers) - set(df_filtered.columns)
                    if missing_columns:
                        self.log_message(f"警告：文件缺少以下列：{', '.join(missing_columns)}")
                        continue
                    
                    # 处理收货日期
                    if '收货日期' in df_filtered.columns:
                        df_filtered['收货日期'] = pd.to_datetime(df_filtered['收货日期'], errors='coerce').dt.strftime('%Y-%m-%d')
                        earliest_date = df_filtered['收货日期'].min()
                        year_month = datetime.strptime(earliest_date, '%Y-%m-%d').strftime('%Y-%m') if earliest_date else None
                        
                        if not year_month:
                            self.log_message("警告：文件中没有有效的收货日期，无法确定年月。")
                            continue
                            
                        # 创建年月子文件夹
                        year_month_folder = os.path.join(output_folder, year_month)
                        if not os.path.exists(year_month_folder):
                            os.makedirs(year_month_folder)
                            
                    # 分组处理
                    group_columns = ['供应商/备用金报销账户', '税率']
                    sort_columns = ['收货日期']
                    
                    # 将税率转换为百分比格式
                    def convert_tax_rate(x):
                        if pd.isna(x):
                            return '0%'
                        tax_str = str(x).strip().replace('%', '')
                        try:
                            tax_value = float(tax_str)
                            if tax_value < 1:  # 如果是小数形式（如0.13）
                                return f"{int(tax_value * 100)}%"
                            return f"{int(tax_value)}%"
                        except ValueError:
                            return '0%'
                    
                    df_filtered['税率'] = df_filtered['税率'].apply(convert_tax_rate)
                    
                    if all(col in df_filtered.columns for col in sort_columns):
                        sorted_df = df_filtered.sort_values(by=sort_columns).groupby(group_columns)
                    else:
                        self.log_message("警告：文件中缺少排序所需的列，将不按顺序处理数据。")
                        sorted_df = df_filtered.groupby(group_columns)
                        
                    # 处理每个分组
                    for group_name, group_data in sorted_df:
                        supplier_account, tax_rate = group_name
                        self.process_group_data(supplier_account, group_data, year_month, year_month_folder, header_rows, tax_rate)
                        
                    # 归档文件
                    archive_filepath = os.path.join(archive_folder, os.path.basename(input_file))
                    if os.path.exists(archive_filepath):
                        base, ext = os.path.splitext(os.path.basename(input_file))
                        timestamp = datetime.now().strftime("%Y%m%d%H%M%S")
                        archive_filepath = os.path.join(archive_folder, f"{base}_{timestamp}{ext}")
                        
                    shutil.move(input_file, archive_filepath)
                    self.log_message(f"已成功归档文件 {os.path.basename(input_file)}")
                    
                    # 更新进度
                    processed_files += 1
                    progress_value = int((processed_files / total_files) * 100)
                    self.progress['value'] = progress_value
                    self.root.update_idletasks()  # 强制更新UI
                    
                except Exception as e:
                    self.log_message(f"处理文件 {os.path.basename(input_file)} 时出错: {str(e)}")
                    
            # 在处理完成后只显示警告信息
            warning_messages = [msg for msg in self.log_messages if msg.startswith("警告：")]
            if warning_messages:
                self.log_message("\n所有文件处理完成。以下是处理过程中的警告信息：")
                for msg in warning_messages:
                    self.log_text.config(state=NORMAL)
                    self.log_text.tag_config("warning", foreground="red")
                    self.log_text.insert(END, msg + "\n", "warning")
                    self.log_text.see(END)
                    self.log_text.config(state=DISABLED)
            else:
                self.log_message("\n所有文件处理完成，没有发现警告信息。")
            
            self.progress['value'] = 100
            
            # 询问是否打开输出目录
            if input_files:
                open_folder = messagebox.askyesno("处理完成", "所有文件处理已完成，是否打开输出文件夹？")
                if open_folder:
                    try:
                        os.startfile(output_folder)
                    except:
                        try:
                            if sys.platform == "darwin":  # macOS
                                subprocess.call(["open", output_folder])
                            else:  # Linux
                                subprocess.call(["xdg-open", output_folder])
                        except:
                            self.log_message("无法打开文件夹，请手动访问：")
                            self.log_message(output_folder)
            
        except Exception as e:
            self.log_message(f"处理过程中发生错误: {str(e)}")
        finally:
            self.processing = False
            self.process_btn.config(state=NORMAL)
            
    def process_group_data(self, group_name, group_data, year_month, year_month_folder, header_rows, tax_rate):
        """处理每个分组的数据"""
        supplier_account = group_name
        
        # 定义红色文字格式
        RED = '\033[91m'
        RESET = '\033[0m'
        
        # 将收货日期转换为datetime格式
        try:
            group_data['收货日期'] = pd.to_datetime(group_data['收货日期'], errors='coerce').dt.strftime('%Y-%m-%d')
            # 检查收货日期是否跨月
            receipt_dates = pd.to_datetime(group_data['收货日期']).dt.strftime('%Y-%m')
            unique_months = receipt_dates.unique()
            if len(unique_months) > 1:
                warning_msg = f"警告：供应商 {supplier_account} 的收货日期包含跨月数据，请核查。包含的月份有：{', '.join(unique_months)}"
                self.log_message(warning_msg)
        except Exception as e:
            error_msg = f"无法解析收货日期：{e}"
            self.log_message(error_msg)
        
        # 构建文件名
        sanitized_supplier_account = ''.join([c if c.isalnum() or c in (' ', '.') else '_' for c in str(supplier_account)])
        sanitized_supplier_account = sanitized_supplier_account.strip('_')
        output_filename = '_'.join(filter(None, [year_month, sanitized_supplier_account, tax_rate])) + '.xlsx'
        output_filepath = os.path.join(year_month_folder, output_filename)
        
        # 创建Excel文件
        wb = Workbook()
        ws = wb.active
        ws.title = "Statement"
        
        # 插入header
        for row in header_rows:
            ws.append(row)
            
        # 写入表头和数据
        ws.append(self.expected_headers)
        yellow_fill = PatternFill(start_color='FFFF00', end_color='FFFF00', fill_type='solid')
        
        # 创建一个列表来存储所有行的数据，包括退货行
        all_rows = []
        subtotal_amount = 0
        tax_amount = 0
        total_amount = 0
        
        for row in dataframe_to_rows(group_data, index=False, header=False):
            # 只保留expected_headers中定义的列
            formatted_row = [row[group_data.columns.get_loc(col)] if col in group_data.columns else '' for col in self.expected_headers]
            
            # 处理税率格式
            tax_rate_idx = self.expected_headers.index('税率')
            if len(formatted_row) > tax_rate_idx:
                tax_rate_value = formatted_row[tax_rate_idx]
                if pd.notna(tax_rate_value):
                    # 税率值已经是百分比格式，直接使用
                    formatted_row[tax_rate_idx] = str(tax_rate_value)
                else:
                    formatted_row[tax_rate_idx] = '0%'
            
            # 添加原始行
            all_rows.append((formatted_row, False))  # False表示不是退货行
            
            # 更新合计金额
            subtotal_idx = self.expected_headers.index('小计金额(结算)')
            tax_idx = self.expected_headers.index('税额(结算)')
            total_idx = self.expected_headers.index('小计价税(结算)')
            
            if len(formatted_row) > max(subtotal_idx, tax_idx, total_idx):
                if pd.notna(formatted_row[subtotal_idx]):
                    subtotal_amount += float(formatted_row[subtotal_idx])
                if pd.notna(formatted_row[tax_idx]):
                    tax_amount += float(formatted_row[tax_idx])
                if pd.notna(formatted_row[total_idx]):
                    total_amount += float(formatted_row[total_idx])
            
            # 检查是否为退货数据并创建退货行
            if '退货' in group_data.columns and pd.notna(row[group_data.columns.get_loc('退货')]) and row[group_data.columns.get_loc('退货')] == '是':
                return_row = formatted_row.copy()
                
                # 设置退货数量（负数）
                if '合计退货数量' in group_data.columns:
                    return_value = -float(row[group_data.columns.get_loc('合计退货数量')])
                    return_row[self.expected_headers.index('实收数量')] = return_value
                
                # 设置退货金额（负数）
                if '退货合计金额(结算)' in group_data.columns:
                    return_value = -float(row[group_data.columns.get_loc('退货合计金额(结算)')])
                    return_row[self.expected_headers.index('小计金额(结算)')] = return_value
                    subtotal_amount += return_value
                
                # 设置退货税额（负数）
                if '退货合计税额(结算)' in group_data.columns:
                    return_value = -float(row[group_data.columns.get_loc('退货合计税额(结算)')])
                    return_row[self.expected_headers.index('税额(结算)')] = return_value
                    tax_amount += return_value
                
                # 设置退货价税合计（负数）
                if '退货合计价税(结算)' in group_data.columns:
                    return_value = -float(row[group_data.columns.get_loc('退货合计价税(结算)')])
                    return_row[self.expected_headers.index('小计价税(结算)')] = return_value
                    total_amount += return_value
                
                all_rows.append((return_row, True))  # True表示是退货行
        
        # 写入所有行并设置样式
        for row_data, is_return in all_rows:
            ws.append(row_data)
            if is_return:  # 如果是退货行，设置黄色背景
                for cell in ws[ws.max_row]:
                    cell.fill = yellow_fill
        
        # 添加合计行
        last_row = ws.max_row + 1
        ws.cell(row=last_row, column=self.expected_headers.index("单价(结算)") + 1, value="合计")
        ws.cell(row=last_row, column=self.expected_headers.index("小计金额(结算)") + 1, value="{:.2f}".format(subtotal_amount))
        ws.cell(row=last_row, column=self.expected_headers.index("税额(结算)") + 1, value="{:.2f}".format(tax_amount))
        ws.cell(row=last_row, column=self.expected_headers.index("小计价税(结算)") + 1, value="{:.2f}".format(total_amount))
        
        # 设置样式
        self.apply_styles(ws)
        
        # 保存文件
        wb.save(output_filepath)
        self.log_message(f"已成功创建 {output_filename}")
        
    def apply_styles(self, ws):
        """应用样式到工作表"""
        # 自动调整列宽
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if cell.row >= 6 and (cell.value is not None and len(str(cell.value)) > max_length):
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 8)
            ws.column_dimensions[column_letter].width = adjusted_width
            
        # 设置页面布局
        ws.page_setup.paperSize = ws.PAPERSIZE_A4
        ws.page_margins = PageMargins(top=0.25, left=0.2, right=0, bottom=1.05, header=0, footer=0.5)
        ws.page_setup.horizontalCentered = True
        ws.page_setup.verticalCentered = True
        ws.sheet_properties.pageSetUpPr.fitToPage = True
        ws.page_setup.fitToHeight = False
        ws.page_setup.fitToWidth = 1
        ws.oddFooter.center.text = "Page &[Page] of &[Pages]"
        ws.print_title_rows = '1:6'
        ws.freeze_panes = 'A7'
        
        # 设置单元格样式
        for row in ws.iter_rows(min_row=1, max_col=len(self.expected_headers), max_row=ws.max_row):
            for cell in row:
                cell.alignment = Alignment(horizontal="center", vertical="center")
                if cell.row <= 5:
                    cell.fill = PatternFill(start_color='1F497D', end_color='1F497D', fill_type='solid')
                    cell.font = Font(color='FFFFFF', size=18, name='微软雅黑', bold=True)
                elif cell.row == 6:
                    cell.fill = PatternFill(start_color='1F497D', end_color='1F497D', fill_type='solid')
                    cell.font = Font(color='FFFFFF', size=10, name='微软雅黑', bold=True)
                elif cell.row == ws.max_row:
                    cell.fill = PatternFill(start_color='1F497D', end_color='1F497D', fill_type='solid')
                    cell.font = Font(color='FFFFFF', size=10, name='微软雅黑', bold=True)
                else:
                    cell.font = Font(size=11, name='微软雅黑')
                    
    def bring_to_front(self):
        """将窗口带到前台"""
        self.root.lift()
        self.root.attributes('-topmost', True)
        self.root.after_idle(self.root.attributes, '-topmost', False)
        
    def create_developer_label(self):
        """在窗口底部创建开发者信息标签"""
        developer_frame = ttk.Frame(self.main_frame)
        developer_frame.pack(side=BOTTOM, fill=X, pady=5)
        
        developer_label = ttk.Label(
            developer_frame,
            text="Powered By Cayman Fu @ Sofitel HAIKOU 2025 Ver 2.3.1",
            font=("微软雅黑", 8),
            foreground="gray"
        )
        developer_label.pack(side=BOTTOM, pady=5)
        
if __name__ == "__main__":
    root = Tk()
    app = BldBuyApp(root)
    root.mainloop()
